#! python3
# excel_worksheet_dictionary.py

import openpyxl, os
from openpyxl.utils import get_column_letter

os.chdir(r'C:\\Users\Niels\Desktop\PEC VR 2')

#make lists for data sets
var1 = []
var2 = []

#make lists to store differences between data sets
listRemoved = []
listNew = []

#open final xlworkbook in cwd folder
filenameList = os.listdir(os.getcwd())
wb = openpyxl.load_workbook(filenameList[-1])

#get worksheet objects for first two worksheets
sheet = wb.worksheets[0]
sheet2 = wb.worksheets[1]

#get highest rows for first two worksheets
highestRow = sheet.max_row
highestRow2 = sheet2.max_row

#store data from cell objects in first worksheet in a list
for i in range(6, highestRow + 1):
    value = sheet.cell(row=i, column=3).value
    var1.append(value)

#store data from cell objects in second worksheet in a list
for i in range(6, highestRow2 + 1):
    value = sheet2.cell(row=i, column=3).value
    var2.append(value)

#create unique set of lists to calculate difference between var1 and var2
listAdded = set(var1) - set(var2)
listAdded = list(listAdded)
listRemoved = set(var2) - set(var1)
listRemoved = list(listRemoved)

#create new worksheets for old and removed pigeons
wb.create_sheet(index=2, title='toegevoegd')
sheetAdded = wb.get_sheet_by_name('toegevoegd')

x = 0
for i in range(1, len(listAdded)+1):
    sheetAdded['A' + str(i)] = listAdded[x]
    x += 1

wb.create_sheet(index=3, title='verwijderd')
sheetRemoved = wb.get_sheet_by_name('verwijderd')

x = 0
for i in range(1, len(listRemoved)+1):
    sheetRemoved['A' + str(i)] = listRemoved[x]
    x += 1

#create dictionary for pigeons that have been removed from the stock
supplyRange = {}

for row in range(6, sheet2.max_row - 7):
    if sheet2['C' + str(row)].value in listRemoved:
        bandnumber = sheet2['C' + str(row)].value
        notes = sheet2['V' + str(row)].value

        supplyRange.setdefault(bandnumber, notes)

#print the notes of the removed pigeons
x = 1
for k, v in supplyRange.items():
    sheetRemoved.cell(column=1, row=x, value=k.format(get_column_letter(1)))
    sheetRemoved.cell(column=2, row=x, value=v.format(get_column_letter(2)))
    x += 1
